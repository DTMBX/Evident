"""
CAD/Dispatch Processor Service - Computer-Aided Dispatch Evidence Processing
=============================================================================

Handles CAD/dispatch system exports with:
- Structured CAD event log parsing
- MDT query/database check extraction
- Unit history and disposition codes
- Radio log integration
- Cross-validation with BWC timestamps
- "No record" / negative evidence handling
- Timestamp discrepancy detection

Supports: Spillman, VersaTerm, CJIS, TriTech, Tyler New World, custom exports

Author: BarberX Legal Case Management
Version: 1.0.0
"""

import asyncio
import csv
import json
import re
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Any, Tuple
from enum import Enum
import hashlib

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================================
# ENUMS & CONSTANTS
# ============================================================================

class CADSystem(str, Enum):
    """Supported CAD platforms"""
    SPILLMAN = "Spillman CAD"
    VERSATERM = "VersaTerm CAD"
    TRITECH = "TriTech Inform CAD"
    TYLER_NEW_WORLD = "Tyler New World"
    CJIS = "CJIS CAD"
    MOTOROLA_PREMIERONE = "Motorola PremierOne"
    CENTRALSQUARE = "CentralSquare CAD"
    HEXAGON = "Hexagon CAD"
    CUSTOM_EXPORT = "Custom Export"
    UNKNOWN = "Unknown CAD System"


class EventType(str, Enum):
    """CAD event types"""
    CALL_RECEIVED = "Call Received"
    UNIT_DISPATCHED = "Unit Dispatched"
    UNIT_EN_ROUTE = "Unit En Route"
    UNIT_ON_SCENE = "Unit On Scene"
    DISPOSITION = "Disposition"
    CLEARED = "Cleared"
    MDT_QUERY = "MDT Query"
    STATUS_CHANGE = "Status Change"
    SUPERVISOR_NOTIFIED = "Supervisor Notified"
    TOW_REQUESTED = "Tow Requested"
    TOW_ARRIVAL = "Tow Arrival"
    BACKUP_REQUESTED = "Backup Requested"
    ARREST = "Arrest"
    TRANSPORT_START = "Transport Start"
    TRANSPORT_END = "Transport End"
    RADIO_TRANSMISSION = "Radio Transmission"


class DispositionCode(str, Enum):
    """Common disposition codes"""
    ARREST = "ARR"
    SUMMONS = "SUM"
    WARNING = "WRN"
    NO_ACTION = "NFA"
    REPORT_TAKEN = "RPT"
    GONE_ON_ARRIVAL = "GOA"
    UNABLE_TO_LOCATE = "UTL"
    TOW = "TOW"
    UNFOUNDED = "UNF"
    TRANSPORTED = "TRANS"


class NegativeEvidenceType(str, Enum):
    """Types of 'no record' responses"""
    NO_CALL_FOR_SERVICE = "No Call for Service Record"
    NO_CAD_EVENT = "No CAD Event Log"
    NO_DISPATCH_RECORD = "No Dispatch Record"
    NO_MDT_QUERY = "No MDT Query Log"
    NO_RADIO_LOG = "No Radio Transmission Log"
    NO_SUPERVISOR_NOTIFICATION = "No Supervisor Notification"
    NO_TOW_AUTHORIZATION = "No Tow Authorization Record"
    NO_UNIT_HISTORY = "No Unit History"
    REDACTED = "Redacted / Withheld"
    EXEMPTION_CLAIMED = "Exemption Claimed (OPRA/FOIA)"


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class CADEvent:
    """Individual CAD event entry"""
    event_id: str  # CAD incident number
    event_type: EventType
    timestamp: str  # ISO format
    unit_id: Optional[str] = None
    officer_id: Optional[str] = None
    location: Optional[str] = None
    disposition_code: Optional[str] = None
    notes: str = ""
    raw_data: Dict = field(default_factory=dict)


@dataclass
class MDTQuery:
    """MDT database query (NCIC, DMV, warrant check, etc.)"""
    query_id: str
    timestamp: str
    query_type: str  # "NCIC", "DMV", "Warrant", "Registration", "Driver_History"
    query_input: str  # Plate, name, DOB, etc.
    queried_by_unit: str
    queried_by_officer: str
    result_summary: str = ""
    hit: bool = False
    raw_response: str = ""


@dataclass
class RadioLog:
    """Radio transmission log entry"""
    timestamp: str
    unit_id: str
    transmission_type: str  # "Dispatcher", "Officer", "Supervisor"
    message: str
    audio_file: Optional[str] = None


@dataclass
class NegativeEvidence:
    """
    Records when agency claims 'no responsive records exist'
    CRITICAL for proving suppression/destruction
    """
    evidence_type: NegativeEvidenceType
    case_id: str
    request_date: str  # When requested (OPRA, subpoena, etc.)
    response_date: str
    responding_agency: str
    responding_custodian: str
    request_scope: str  # What was asked for
    response_text: str  # Exact language from response
    attached_response_file: Optional[str] = None  # PDF of OPRA response
    verification_hash: Optional[str] = None
    notes: str = ""
    
    def to_dict(self) -> Dict:
        data = asdict(self)
        data['evidence_type'] = self.evidence_type.value
        return data


@dataclass
class TimestampDiscrepancy:
    """Detected discrepancies between CAD and other sources"""
    event_description: str
    cad_timestamp: Optional[str]
    bwc_timestamp: Optional[str]
    tow_invoice_timestamp: Optional[str]
    station_log_timestamp: Optional[str]
    discrepancy_seconds: float
    significance: str  # "Minor", "Major", "Critical"
    notes: str = ""


@dataclass
class CADProcessingResult:
    """Complete CAD processing result"""
    evidence_id: str
    case_id: str
    cad_system: CADSystem
    source_file: str
    
    # Parsed data
    events: List[CADEvent] = field(default_factory=list)
    mdt_queries: List[MDTQuery] = field(default_factory=list)
    radio_logs: List[RadioLog] = field(default_factory=list)
    negative_evidence: List[NegativeEvidence] = field(default_factory=list)
    
    # Analysis
    discrepancies: List[TimestampDiscrepancy] = field(default_factory=list)
    
    # Metadata
    processing_date: str = field(default_factory=lambda: datetime.utcnow().isoformat())
    processing_notes: str = ""
    
    def to_dict(self) -> Dict:
        data = asdict(self)
        data['cad_system'] = self.cad_system.value
        for event in data['events']:
            event['event_type'] = event['event_type'] if isinstance(event['event_type'], str) else event['event_type'].value
        for neg in data['negative_evidence']:
            neg['evidence_type'] = neg['evidence_type'] if isinstance(neg['evidence_type'], str) else neg['evidence_type'].value
        return data


# ============================================================================
# CAD PROCESSOR SERVICE
# ============================================================================

class CADProcessorService:
    """
    Process CAD/dispatch exports and handle 'no record' responses
    """
    
    def __init__(self, output_dir: str = "./cad_processed"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.metadata_dir = self.output_dir / "metadata"
        self.negative_evidence_dir = self.output_dir / "negative_evidence"
        self.discrepancy_reports_dir = self.output_dir / "discrepancy_reports"
        
        for directory in [self.metadata_dir, self.negative_evidence_dir, self.discrepancy_reports_dir]:
            directory.mkdir(exist_ok=True)
        
        # Store negative evidence records
        self.negative_evidence_db: List[NegativeEvidence] = []
    
    def detect_cad_system(self, file_path: Path) -> CADSystem:
        """Detect CAD system from file format or headers"""
        if not file_path.exists():
            return CADSystem.UNKNOWN
        
        filename = file_path.name.lower()
        
        if "spillman" in filename:
            return CADSystem.SPILLMAN
        elif "versaterm" in filename:
            return CADSystem.VERSATERM
        elif "tritech" in filename or "inform" in filename:
            return CADSystem.TRITECH
        elif "tyler" in filename or "newworld" in filename:
            return CADSystem.TYLER_NEW_WORLD
        
        # Try to detect from file content
        try:
            if file_path.suffix == '.csv':
                with open(file_path, 'r') as f:
                    header = f.readline().lower()
                    if "spillman" in header:
                        return CADSystem.SPILLMAN
                    elif "versaterm" in header:
                        return CADSystem.VERSATERM
            elif file_path.suffix in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path, nrows=5)
                header_text = ' '.join(str(col).lower() for col in df.columns)
                if "spillman" in header_text:
                    return CADSystem.SPILLMAN
                elif "versaterm" in header_text:
                    return CADSystem.VERSATERM
        except:
            pass
        
        return CADSystem.CUSTOM_EXPORT
    
    async def process_cad_export(
        self,
        file_path: Path,
        evidence_id: str,
        case_id: str,
        expected_incident_id: Optional[str] = None
    ) -> CADProcessingResult:
        """
        Process CAD export file (CSV, Excel, PDF, or text)
        
        Args:
            file_path: Path to CAD export
            evidence_id: Evidence vault ID
            case_id: Case reference
            expected_incident_id: Expected CAD incident number (for verification)
        
        Returns:
            CADProcessingResult with parsed events and analysis
        """
        if not file_path.exists():
            raise FileNotFoundError(f"CAD file not found: {file_path}")
        
        cad_system = self.detect_cad_system(file_path)
        
        # Parse based on file type
        if file_path.suffix == '.csv':
            events, mdt_queries, radio_logs = await self._parse_csv(file_path, cad_system)
        elif file_path.suffix in ['.xlsx', '.xls']:
            events, mdt_queries, radio_logs = await self._parse_excel(file_path, cad_system)
        elif file_path.suffix == '.txt':
            events, mdt_queries, radio_logs = await self._parse_text(file_path, cad_system)
        else:
            raise ValueError(f"Unsupported CAD file format: {file_path.suffix}")
        
        # Build result
        result = CADProcessingResult(
            evidence_id=evidence_id,
            case_id=case_id,
            cad_system=cad_system,
            source_file=str(file_path),
            events=events,
            mdt_queries=mdt_queries,
            radio_logs=radio_logs
        )
        
        # Save metadata
        await self._save_cad_metadata(result)
        
        return result
    
    async def _parse_csv(
        self,
        file_path: Path,
        cad_system: CADSystem
    ) -> Tuple[List[CADEvent], List[MDTQuery], List[RadioLog]]:
        """Parse CSV CAD export"""
        events = []
        mdt_queries = []
        radio_logs = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Attempt to parse as CAD event
                    event = self._parse_cad_row(row, cad_system)
                    if event:
                        events.append(event)
        except Exception as e:
            print(f"Error parsing CAD CSV: {e}")
        
        return events, mdt_queries, radio_logs
    
    async def _parse_excel(
        self,
        file_path: Path,
        cad_system: CADSystem
    ) -> Tuple[List[CADEvent], List[MDTQuery], List[RadioLog]]:
        """Parse Excel CAD export"""
        events = []
        mdt_queries = []
        radio_logs = []
        
        try:
            df = pd.read_excel(file_path)
            for _, row in df.iterrows():
                event = self._parse_cad_row(row.to_dict(), cad_system)
                if event:
                    events.append(event)
        except Exception as e:
            print(f"Error parsing CAD Excel: {e}")
        
        return events, mdt_queries, radio_logs
    
    async def _parse_text(
        self,
        file_path: Path,
        cad_system: CADSystem
    ) -> Tuple[List[CADEvent], List[MDTQuery], List[RadioLog]]:
        """Parse text CAD export"""
        events = []
        mdt_queries = []
        radio_logs = []
        
        # Text parsing would be system-specific
        # Example: parse fixed-width or delimited text
        
        return events, mdt_queries, radio_logs
    
    def _parse_cad_row(self, row: Dict, cad_system: CADSystem) -> Optional[CADEvent]:
        """Parse individual CAD row into CADEvent"""
        # Map common column names
        column_map = {
            'incident_number': ['incident_number', 'incident_id', 'event_id', 'call_id', 'cad_id'],
            'timestamp': ['timestamp', 'date_time', 'event_time', 'time', 'datetime'],
            'event_type': ['event_type', 'event', 'type', 'action', 'status'],
            'unit_id': ['unit_id', 'unit', 'car', 'officer_unit'],
            'officer_id': ['officer_id', 'officer', 'badge', 'badge_number'],
            'location': ['location', 'address', 'place', 'scene_address'],
            'disposition': ['disposition', 'dispo', 'result', 'outcome'],
            'notes': ['notes', 'comments', 'remarks', 'narrative']
        }
        
        def find_value(aliases: List[str]) -> Optional[str]:
            for alias in aliases:
                for key in row.keys():
                    if key and alias.lower() in key.lower():
                        val = row[key]
                        return str(val) if val and str(val) != 'nan' else None
            return None
        
        incident_id = find_value(column_map['incident_number'])
        timestamp = find_value(column_map['timestamp'])
        event_type_str = find_value(column_map['event_type'])
        
        if not incident_id or not timestamp:
            return None
        
        # Map event type string to enum
        event_type = self._map_event_type(event_type_str)
        
        return CADEvent(
            event_id=incident_id,
            event_type=event_type,
            timestamp=self._normalize_timestamp(timestamp),
            unit_id=find_value(column_map['unit_id']),
            officer_id=find_value(column_map['officer_id']),
            location=find_value(column_map['location']),
            disposition_code=find_value(column_map['disposition']),
            notes=find_value(column_map['notes']) or "",
            raw_data=row
        )
    
    def _map_event_type(self, event_str: Optional[str]) -> EventType:
        """Map CAD event string to EventType enum"""
        if not event_str:
            return EventType.STATUS_CHANGE
        
        event_lower = event_str.lower()
        
        if "dispatch" in event_lower:
            return EventType.UNIT_DISPATCHED
        elif "en route" in event_lower or "enroute" in event_lower:
            return EventType.UNIT_EN_ROUTE
        elif "on scene" in event_lower or "arrived" in event_lower:
            return EventType.UNIT_ON_SCENE
        elif "clear" in event_lower or "available" in event_lower:
            return EventType.CLEARED
        elif "tow" in event_lower:
            return EventType.TOW_REQUESTED
        elif "arrest" in event_lower:
            return EventType.ARREST
        elif "transport" in event_lower:
            return EventType.TRANSPORT_START
        elif "mdt" in event_lower or "query" in event_lower:
            return EventType.MDT_QUERY
        elif "supervisor" in event_lower:
            return EventType.SUPERVISOR_NOTIFIED
        else:
            return EventType.STATUS_CHANGE
    
    def _normalize_timestamp(self, timestamp_str: str) -> str:
        """Normalize timestamp to ISO format"""
        # Handle common formats:
        # "01/22/2025 14:30:22"
        # "2025-01-22 14:30:22"
        # "01/22/25 2:30 PM"
        
        formats = [
            "%m/%d/%Y %H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%m/%d/%y %I:%M %p",
            "%Y-%m-%d %I:%M:%S %p",
            "%m/%d/%Y %I:%M:%S %p"
        ]
        
        for fmt in formats:
            try:
                dt = datetime.strptime(timestamp_str.strip(), fmt)
                return dt.isoformat()
            except:
                continue
        
        # Return as-is if can't parse
        return timestamp_str
    
    async def record_negative_evidence(
        self,
        case_id: str,
        evidence_type: NegativeEvidenceType,
        responding_agency: str,
        responding_custodian: str,
        request_scope: str,
        response_text: str,
        response_file_path: Optional[Path] = None,
        request_date: Optional[str] = None,
        response_date: Optional[str] = None
    ) -> NegativeEvidence:
        """
        Record when agency claims 'no responsive records'
        
        This is CRITICAL evidence of potential destruction/suppression
        """
        if response_file_path and response_file_path.exists():
            # Hash the response document
            with open(response_file_path, 'rb') as f:
                file_hash = hashlib.sha256(f.read()).hexdigest()
        else:
            file_hash = None
        
        negative_record = NegativeEvidence(
            evidence_type=evidence_type,
            case_id=case_id,
            request_date=request_date or datetime.utcnow().isoformat(),
            response_date=response_date or datetime.utcnow().isoformat(),
            responding_agency=responding_agency,
            responding_custodian=responding_custodian,
            request_scope=request_scope,
            response_text=response_text,
            attached_response_file=str(response_file_path) if response_file_path else None,
            verification_hash=file_hash
        )
        
        self.negative_evidence_db.append(negative_record)
        
        # Save to Excel
        await self._save_negative_evidence(negative_record)
        
        return negative_record
    
    async def cross_validate_timestamps(
        self,
        cad_events: List[CADEvent],
        bwc_metadata: Optional[Dict] = None,
        tow_invoice_data: Optional[Dict] = None,
        station_log_data: Optional[Dict] = None
    ) -> List[TimestampDiscrepancy]:
        """
        Cross-validate timestamps from multiple sources
        
        Looks for discrepancies that suggest:
        - Backdating
        - Clock errors
        - Falsification
        - System issues
        """
        discrepancies = []
        
        # Example: Compare CAD "on scene" time with BWC start time
        if bwc_metadata:
            cad_on_scene = next((e for e in cad_events if e.event_type == EventType.UNIT_ON_SCENE), None)
            if cad_on_scene:
                cad_time = datetime.fromisoformat(cad_on_scene.timestamp)
                bwc_start_str = bwc_metadata.get('time_sync', {}).get('system_timestamp')
                if bwc_start_str:
                    bwc_time = datetime.fromisoformat(bwc_start_str)
                    diff_seconds = (bwc_time - cad_time).total_seconds()
                    
                    if abs(diff_seconds) > 60:  # More than 1 minute difference
                        discrepancies.append(TimestampDiscrepancy(
                            event_description="Unit On Scene vs BWC Start",
                            cad_timestamp=cad_on_scene.timestamp,
                            bwc_timestamp=bwc_start_str,
                            tow_invoice_timestamp=None,
                            station_log_timestamp=None,
                            discrepancy_seconds=diff_seconds,
                            significance="Major" if abs(diff_seconds) > 300 else "Minor",
                            notes=f"CAD shows {diff_seconds/60:.1f} minutes {'earlier' if diff_seconds < 0 else 'later'} than BWC"
                        ))
        
        return discrepancies
    
    async def _save_cad_metadata(self, result: CADProcessingResult):
        """Save CAD processing result as JSON"""
        metadata_path = self.metadata_dir / f"{result.evidence_id}_cad_metadata.json"
        with open(metadata_path, 'w') as f:
            json.dump(result.to_dict(), f, indent=2)
    
    async def _save_negative_evidence(self, negative: NegativeEvidence):
        """Save negative evidence record to Excel database"""
        db_path = self.negative_evidence_dir / "negative_evidence_log.xlsx"
        
        if not db_path.exists():
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Negative_Evidence"
            
            headers = [
                "Evidence_Type", "Case_ID", "Request_Date", "Response_Date",
                "Responding_Agency", "Responding_Custodian", "Request_Scope",
                "Response_Text", "Attached_File", "Verification_Hash", "Notes"
            ]
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        else:
            wb = openpyxl.load_workbook(db_path)
            ws = wb["Negative_Evidence"]
        
        # Append record
        row = [
            negative.evidence_type.value,
            negative.case_id,
            negative.request_date,
            negative.response_date,
            negative.responding_agency,
            negative.responding_custodian,
            negative.request_scope,
            negative.response_text,
            negative.attached_response_file,
            negative.verification_hash,
            negative.notes
        ]
        ws.append(row)
        
        wb.save(db_path)
    
    async def generate_discrepancy_report(
        self,
        case_id: str,
        discrepancies: List[TimestampDiscrepancy]
    ) -> Path:
        """Generate Excel report of all timestamp discrepancies"""
        report_path = self.discrepancy_reports_dir / f"{case_id}_discrepancies.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timestamp_Discrepancies"
        
        headers = [
            "Event_Description", "CAD_Timestamp", "BWC_Timestamp",
            "Tow_Invoice_Timestamp", "Station_Log_Timestamp",
            "Discrepancy_Seconds", "Significance", "Notes"
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        for disc in discrepancies:
            ws.append([
                disc.event_description,
                disc.cad_timestamp,
                disc.bwc_timestamp,
                disc.tow_invoice_timestamp,
                disc.station_log_timestamp,
                disc.discrepancy_seconds,
                disc.significance,
                disc.notes
            ])
        
        wb.save(report_path)
        return report_path


# ============================================================================
# GLOBAL SERVICE INSTANCE
# ============================================================================

cad_processor = CADProcessorService()

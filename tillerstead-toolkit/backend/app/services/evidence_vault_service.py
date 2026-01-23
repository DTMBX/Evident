"""
Evidence Vault Service - Defensible Collection & Chain of Custody
==================================================================

Implements court-defensible evidence management with:
- SHA-256 hashing at ingestion
- Immutable/WORM-capable storage
- Role-based access control
- Complete chain-of-custody logging
- Litigation hold management
- Provenance tracking

Author: BarberX Legal Case Management
Version: 1.0.0
"""

import hashlib
import json
import os
import shutil
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from enum import Enum
from pathlib import Path
from typing import Dict, List, Optional, Set, Any
import asyncio
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd


# ============================================================================
# ENUMS & CONSTANTS
# ============================================================================

class EvidenceType(str, Enum):
    """Evidence media types"""
    BWC_VIDEO = "BWC Video"
    DASH_CAM = "Dash Camera"
    AUDIO = "Audio Recording"
    PDF_DOCUMENT = "PDF Document"
    CAD_EXPORT = "CAD/Dispatch Export"
    MDT_LOG = "MDT Query Log"
    TOW_INVOICE = "Tow Invoice"
    OPRA_RESPONSE = "OPRA Response"
    EMAIL = "Email"
    PHOTOGRAPH = "Photograph"
    SPREADSHEET = "Spreadsheet"
    TEXT_MESSAGE = "Text Message"
    DATABASE_EXPORT = "Database Export"
    OTHER = "Other"


class SourceSystem(str, Enum):
    """Source systems for evidence"""
    AXON_EVIDENCE = "Axon Evidence.com"
    WATCHGUARD = "WatchGuard 4RE"
    MOTOROLA = "Motorola BWC"
    VIEVU = "Vievu LE5"
    SPILLMAN = "Spillman CAD"
    VERSATERM = "VersaTerm CAD"
    ACPO_OPRA = "Atlantic County Prosecutor OPRA"
    MUNICIPALITY_OPRA = "Municipal OPRA"
    TOWING_COMPANY = "Towing Company"
    MANUAL_UPLOAD = "Manual Upload"
    EMAIL_EXPORT = "Email Export"
    SUBPOENA_RESPONSE = "Subpoena Response"
    DISCOVERY_PRODUCTION = "Discovery Production"
    OTHER = "Other Source"


class ExportMethod(str, Enum):
    """How evidence was obtained"""
    NATIVE_EXPORT = "Native Export"
    API_EXPORT = "API Export"
    SCREEN_CAPTURE = "Screen Capture"
    PHOTOGRAPH = "Photograph"
    MANUAL_DOWNLOAD = "Manual Download"
    EMAIL_ATTACHMENT = "Email Attachment"
    PHYSICAL_MEDIA = "Physical Media (CD/DVD/USB)"
    FTP_TRANSFER = "FTP Transfer"
    OPRA_DISCLOSURE = "OPRA Disclosure"
    SUBPOENA_PRODUCTION = "Subpoena Production"


class AccessLevel(str, Enum):
    """Role-based access levels"""
    ADMIN = "Administrator"
    ATTORNEY = "Attorney"
    PARALEGAL = "Paralegal"
    INVESTIGATOR = "Investigator"
    EXPERT = "Expert Witness"
    VIEWER = "Read-Only Viewer"
    RESTRICTED = "Restricted (Protective Order)"


class ChainEventType(str, Enum):
    """Chain of custody event types"""
    INGESTION = "Ingestion"
    HASH_VERIFICATION = "Hash Verification"
    ACCESS = "Access"
    EXPORT = "Export"
    PROCESSING = "Processing"
    COPY = "Copy"
    MOVE = "Move"
    LOCK = "Litigation Hold Applied"
    UNLOCK = "Litigation Hold Released"
    DELETION_REQUESTED = "Deletion Requested"
    DELETION_EXECUTED = "Deletion Executed"
    METADATA_CHANGE = "Metadata Updated"
    PERMISSION_CHANGE = "Permission Changed"


class LitigationHoldStatus(str, Enum):
    """Litigation hold states"""
    ACTIVE = "Active Hold"
    RELEASED = "Released"
    PENDING_REVIEW = "Pending Review"
    EXPIRED = "Expired"


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class EvidenceProvenance:
    """Provenance tracking for evidence items"""
    source_system: SourceSystem
    export_method: ExportMethod
    received_from: str  # Person/entity name
    custodian: str  # Original custodian
    collection_date: str  # ISO format
    received_date: str  # ISO format
    original_filename: str
    original_location: str  # Path on source system
    export_credentials: Optional[str] = None  # Who performed export
    chain_custody_notes: str = ""
    verification_method: str = "SHA-256"
    
    def to_dict(self) -> Dict:
        return asdict(self)


@dataclass
class EvidenceItem:
    """Core evidence item with full chain of custody"""
    evidence_id: str  # Unique ID (e.g., "EV-2025-001")
    case_id: str  # Case reference
    evidence_type: EvidenceType
    original_filename: str
    stored_filename: str  # UUID-based storage name
    file_size_bytes: int
    sha256_hash: str
    provenance: EvidenceProvenance
    
    # Storage paths
    vault_path: str  # Immutable original
    working_path: Optional[str] = None  # Processed/working copy
    
    # Metadata
    description: str = ""
    tags: List[str] = field(default_factory=list)
    custodians: List[str] = field(default_factory=list)
    
    # Timestamps
    ingestion_timestamp: str = field(default_factory=lambda: datetime.utcnow().isoformat())
    last_accessed: Optional[str] = None
    last_verified: Optional[str] = None
    
    # Access control
    access_level: AccessLevel = AccessLevel.ATTORNEY
    allowed_users: List[str] = field(default_factory=list)
    
    # Litigation hold
    litigation_hold: bool = False
    litigation_hold_date: Optional[str] = None
    litigation_hold_reason: Optional[str] = None
    litigation_hold_status: LitigationHoldStatus = LitigationHoldStatus.ACTIVE
    
    # Processing
    processed: bool = False
    processing_notes: str = ""
    derived_items: List[str] = field(default_factory=list)  # IDs of derived evidence
    
    def to_dict(self) -> Dict:
        data = asdict(self)
        # Convert enums to strings
        data['evidence_type'] = self.evidence_type.value
        data['access_level'] = self.access_level.value
        data['litigation_hold_status'] = self.litigation_hold_status.value
        data['provenance']['source_system'] = self.provenance.source_system.value
        data['provenance']['export_method'] = self.provenance.export_method.value
        return data


@dataclass
class ChainOfCustodyEvent:
    """Individual chain of custody event"""
    event_id: str
    evidence_id: str
    event_type: ChainEventType
    timestamp: str
    user: str
    user_role: str
    action_description: str
    hash_before: Optional[str] = None
    hash_after: Optional[str] = None
    ip_address: Optional[str] = None
    workstation: Optional[str] = None
    notes: str = ""
    
    def to_dict(self) -> Dict:
        data = asdict(self)
        data['event_type'] = self.event_type.value
        return data


# ============================================================================
# EVIDENCE VAULT SERVICE
# ============================================================================

class EvidenceVaultService:
    """
    Court-defensible evidence vault with immutable storage and complete audit trails
    """
    
    def __init__(self, vault_root: str = "./evidence_vault"):
        self.vault_root = Path(vault_root)
        self.vault_storage = self.vault_root / "vault_storage"  # WORM/immutable
        self.working_storage = self.vault_root / "working_storage"  # Processed copies
        self.exports_storage = self.vault_root / "exports"
        self.database_path = self.vault_root / "evidence_database.xlsx"
        self.chain_log_path = self.vault_root / "chain_of_custody.xlsx"
        self.hash_manifest_path = self.vault_root / "hash_manifests"
        
        # In-memory caches
        self.evidence_items: Dict[str, EvidenceItem] = {}
        self.chain_events: List[ChainOfCustodyEvent] = []
        self.litigation_holds: Dict[str, List[str]] = defaultdict(list)  # case_id -> [evidence_ids]
        
        # Initialize
        self._initialize_vault()
    
    def _initialize_vault(self):
        """Create vault directory structure"""
        self.vault_storage.mkdir(parents=True, exist_ok=True)
        self.working_storage.mkdir(parents=True, exist_ok=True)
        self.exports_storage.mkdir(parents=True, exist_ok=True)
        self.hash_manifest_path.mkdir(parents=True, exist_ok=True)
        
        # Create databases if they don't exist
        if not self.database_path.exists():
            self._create_evidence_database()
        else:
            self._load_evidence_database()
        
        if not self.chain_log_path.exists():
            self._create_chain_log()
        else:
            self._load_chain_log()
    
    def _create_evidence_database(self):
        """Create Excel database for evidence items"""
        wb = openpyxl.Workbook()
        
        # Sheet 1: Evidence Items
        ws = wb.active
        ws.title = "Evidence_Items"
        headers = [
            "Evidence_ID", "Case_ID", "Evidence_Type", "Original_Filename",
            "Stored_Filename", "File_Size_Bytes", "SHA256_Hash",
            "Source_System", "Export_Method", "Received_From", "Custodian",
            "Collection_Date", "Received_Date", "Ingestion_Timestamp",
            "Vault_Path", "Working_Path", "Description", "Tags",
            "Access_Level", "Litigation_Hold", "Hold_Reason", "Hold_Date",
            "Processed", "Processing_Notes"
        ]
        ws.append(headers)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Sheet 2: Litigation Holds
        ws2 = wb.create_sheet("Litigation_Holds")
        ws2.append([
            "Case_ID", "Hold_ID", "Evidence_IDs", "Hold_Date", "Hold_Reason",
            "Status", "Released_Date", "Released_By", "Notes"
        ])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        # Sheet 3: Access Control
        ws3 = wb.create_sheet("Access_Control")
        ws3.append([
            "Evidence_ID", "User_Name", "User_Role", "Access_Level",
            "Granted_Date", "Granted_By", "Expires_Date", "Notes"
        ])
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        
        # Sheet 4: Hash Manifests
        ws4 = wb.create_sheet("Hash_Manifests")
        ws4.append([
            "Manifest_ID", "Date_Created", "Evidence_Count", "Total_Size_GB",
            "Manifest_Hash", "Verified_By", "Verified_Date", "Status"
        ])
        for cell in ws4[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        
        wb.save(self.database_path)
    
    def _create_chain_log(self):
        """Create Excel log for chain of custody events"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chain_Of_Custody"
        
        headers = [
            "Event_ID", "Evidence_ID", "Event_Type", "Timestamp",
            "User", "User_Role", "Action_Description",
            "Hash_Before", "Hash_After", "IP_Address", "Workstation", "Notes"
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        wb.save(self.chain_log_path)
    
    def _load_evidence_database(self):
        """Load evidence items from Excel"""
        try:
            df = pd.read_excel(self.database_path, sheet_name="Evidence_Items")
            for _, row in df.iterrows():
                # Reconstruct EvidenceItem from row
                # (Simplified - full implementation would deserialize all fields)
                evidence_id = row.get("Evidence_ID")
                if evidence_id and pd.notna(evidence_id):
                    self.evidence_items[evidence_id] = row.to_dict()
        except Exception as e:
            print(f"Warning: Could not load evidence database: {e}")
    
    def _load_chain_log(self):
        """Load chain of custody events from Excel"""
        try:
            df = pd.read_excel(self.chain_log_path, sheet_name="Chain_Of_Custody")
            for _, row in df.iterrows():
                # Reconstruct ChainOfCustodyEvent from row
                self.chain_events.append(row.to_dict())
        except Exception as e:
            print(f"Warning: Could not load chain log: {e}")
    
    def _compute_sha256(self, file_path: Path) -> str:
        """Compute SHA-256 hash of file"""
        sha256 = hashlib.sha256()
        with open(file_path, 'rb') as f:
            while chunk := f.read(8192):
                sha256.update(chunk)
        return sha256.hexdigest()
    
    def _generate_evidence_id(self, case_id: str) -> str:
        """Generate unique evidence ID"""
        year = datetime.now().year
        count = len([e for e in self.evidence_items.values() if isinstance(e, dict) and e.get('Case_ID') == case_id])
        return f"EV-{case_id}-{year}-{count+1:04d}"
    
    def _generate_event_id(self) -> str:
        """Generate unique chain event ID"""
        timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
        return f"CHAIN-{timestamp}"
    
    async def ingest_evidence(
        self,
        file_path: Path,
        case_id: str,
        evidence_type: EvidenceType,
        provenance: EvidenceProvenance,
        description: str = "",
        tags: List[str] = None,
        user: str = "System",
        user_role: str = "Administrator"
    ) -> EvidenceItem:
        """
        Ingest new evidence with full chain of custody
        
        Returns:
            EvidenceItem with vault path and hash
        """
        if not file_path.exists():
            raise FileNotFoundError(f"Source file not found: {file_path}")
        
        # Generate evidence ID
        evidence_id = self._generate_evidence_id(case_id)
        
        # Compute hash BEFORE moving to vault
        original_hash = self._compute_sha256(file_path)
        
        # Generate storage filename (preserve extension)
        extension = file_path.suffix
        stored_filename = f"{evidence_id}{extension}"
        vault_path = self.vault_storage / stored_filename
        
        # Copy to vault (immutable storage)
        shutil.copy2(file_path, vault_path)
        
        # Verify hash after copy
        vault_hash = self._compute_sha256(vault_path)
        if original_hash != vault_hash:
            raise ValueError(f"Hash mismatch during ingestion! Original: {original_hash}, Vault: {vault_hash}")
        
        # Make file read-only (simulate WORM)
        os.chmod(vault_path, 0o444)
        
        # Create evidence item
        evidence = EvidenceItem(
            evidence_id=evidence_id,
            case_id=case_id,
            evidence_type=evidence_type,
            original_filename=file_path.name,
            stored_filename=stored_filename,
            file_size_bytes=vault_path.stat().st_size,
            sha256_hash=vault_hash,
            provenance=provenance,
            vault_path=str(vault_path),
            description=description,
            tags=tags or []
        )
        
        # Store in memory
        self.evidence_items[evidence_id] = evidence
        
        # Log chain of custody event
        chain_event = ChainOfCustodyEvent(
            event_id=self._generate_event_id(),
            evidence_id=evidence_id,
            event_type=ChainEventType.INGESTION,
            timestamp=datetime.utcnow().isoformat(),
            user=user,
            user_role=user_role,
            action_description=f"Ingested {evidence_type.value} from {provenance.source_system.value}",
            hash_before=original_hash,
            hash_after=vault_hash,
            notes=f"Original: {file_path.name}"
        )
        self.chain_events.append(chain_event)
        
        # Save to Excel
        self._save_evidence_to_excel(evidence)
        self._save_chain_event_to_excel(chain_event)
        
        # Create hash manifest
        self._create_hash_manifest([evidence])
        
        return evidence
    
    async def verify_evidence_integrity(self, evidence_id: str, user: str = "System") -> bool:
        """
        Verify evidence hash matches original ingestion hash
        
        Returns:
            True if hash matches, False otherwise
        """
        if evidence_id not in self.evidence_items:
            raise ValueError(f"Evidence not found: {evidence_id}")
        
        evidence = self.evidence_items[evidence_id]
        vault_path = Path(evidence.vault_path)
        
        if not vault_path.exists():
            raise FileNotFoundError(f"Vault file missing: {vault_path}")
        
        # Compute current hash
        current_hash = self._compute_sha256(vault_path)
        original_hash = evidence.sha256_hash
        
        # Log verification event
        chain_event = ChainOfCustodyEvent(
            event_id=self._generate_event_id(),
            evidence_id=evidence_id,
            event_type=ChainEventType.HASH_VERIFICATION,
            timestamp=datetime.utcnow().isoformat(),
            user=user,
            user_role="System",
            action_description="Hash integrity verification",
            hash_before=original_hash,
            hash_after=current_hash,
            notes="PASS" if current_hash == original_hash else "FAIL - HASH MISMATCH"
        )
        self.chain_events.append(chain_event)
        self._save_chain_event_to_excel(chain_event)
        
        # Update last verified timestamp
        evidence.last_verified = datetime.utcnow().isoformat()
        self._save_evidence_to_excel(evidence)
        
        return current_hash == original_hash
    
    async def apply_litigation_hold(
        self,
        case_id: str,
        evidence_ids: List[str],
        reason: str,
        user: str,
        user_role: str = "Attorney"
    ):
        """
        Apply litigation hold to evidence items
        """
        hold_date = datetime.utcnow().isoformat()
        
        for evidence_id in evidence_ids:
            if evidence_id not in self.evidence_items:
                continue
            
            evidence = self.evidence_items[evidence_id]
            evidence.litigation_hold = True
            evidence.litigation_hold_date = hold_date
            evidence.litigation_hold_reason = reason
            evidence.litigation_hold_status = LitigationHoldStatus.ACTIVE
            
            # Log chain event
            chain_event = ChainOfCustodyEvent(
                event_id=self._generate_event_id(),
                evidence_id=evidence_id,
                event_type=ChainEventType.LOCK,
                timestamp=hold_date,
                user=user,
                user_role=user_role,
                action_description=f"Litigation hold applied: {reason}",
                notes=f"Case: {case_id}"
            )
            self.chain_events.append(chain_event)
            self._save_chain_event_to_excel(chain_event)
            self._save_evidence_to_excel(evidence)
        
        # Track litigation holds by case
        self.litigation_holds[case_id].extend(evidence_ids)
    
    def _save_evidence_to_excel(self, evidence: EvidenceItem):
        """Append or update evidence item in Excel"""
        wb = openpyxl.load_workbook(self.database_path)
        ws = wb["Evidence_Items"]
        
        row_data = [
            evidence.evidence_id,
            evidence.case_id,
            evidence.evidence_type.value,
            evidence.original_filename,
            evidence.stored_filename,
            evidence.file_size_bytes,
            evidence.sha256_hash,
            evidence.provenance.source_system.value,
            evidence.provenance.export_method.value,
            evidence.provenance.received_from,
            evidence.provenance.custodian,
            evidence.provenance.collection_date,
            evidence.provenance.received_date,
            evidence.ingestion_timestamp,
            evidence.vault_path,
            evidence.working_path,
            evidence.description,
            ", ".join(evidence.tags),
            evidence.access_level.value,
            evidence.litigation_hold,
            evidence.litigation_hold_reason,
            evidence.litigation_hold_date,
            evidence.processed,
            evidence.processing_notes
        ]
        
        ws.append(row_data)
        wb.save(self.database_path)
    
    def _save_chain_event_to_excel(self, event: ChainOfCustodyEvent):
        """Append chain of custody event to Excel"""
        wb = openpyxl.load_workbook(self.chain_log_path)
        ws = wb["Chain_Of_Custody"]
        
        row_data = [
            event.event_id,
            event.evidence_id,
            event.event_type.value,
            event.timestamp,
            event.user,
            event.user_role,
            event.action_description,
            event.hash_before,
            event.hash_after,
            event.ip_address,
            event.workstation,
            event.notes
        ]
        
        ws.append(row_data)
        wb.save(self.chain_log_path)
    
    def _create_hash_manifest(self, evidence_items: List[EvidenceItem]) -> Path:
        """Create hash manifest file for evidence batch"""
        manifest_id = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
        manifest_path = self.hash_manifest_path / f"MANIFEST-{manifest_id}.json"
        
        manifest = {
            "manifest_id": manifest_id,
            "created": datetime.utcnow().isoformat(),
            "evidence_count": len(evidence_items),
            "items": []
        }
        
        for evidence in evidence_items:
            manifest["items"].append({
                "evidence_id": evidence.evidence_id,
                "filename": evidence.original_filename,
                "sha256": evidence.sha256_hash,
                "size_bytes": evidence.file_size_bytes,
                "vault_path": evidence.vault_path
            })
        
        # Compute manifest hash
        manifest_json = json.dumps(manifest, sort_keys=True, indent=2)
        manifest_hash = hashlib.sha256(manifest_json.encode()).hexdigest()
        manifest["manifest_hash"] = manifest_hash
        
        # Save manifest
        with open(manifest_path, 'w') as f:
            json.dump(manifest, f, indent=2)
        
        return manifest_path
    
    async def generate_chain_report(self, evidence_id: str) -> Dict:
        """
        Generate printable chain of custody report for court
        """
        if evidence_id not in self.evidence_items:
            raise ValueError(f"Evidence not found: {evidence_id}")
        
        evidence = self.evidence_items[evidence_id]
        events = [e for e in self.chain_events if e.evidence_id == evidence_id]
        
        report = {
            "evidence_id": evidence_id,
            "case_id": evidence.case_id,
            "evidence_type": evidence.evidence_type.value,
            "original_filename": evidence.original_filename,
            "sha256_hash": evidence.sha256_hash,
            "file_size_mb": round(evidence.file_size_bytes / 1024 / 1024, 2),
            "ingestion_date": evidence.ingestion_timestamp,
            "provenance": evidence.provenance.to_dict(),
            "litigation_hold": evidence.litigation_hold,
            "chain_events": [e.to_dict() for e in sorted(events, key=lambda x: x.timestamp)],
            "total_events": len(events),
            "last_verified": evidence.last_verified,
            "report_generated": datetime.utcnow().isoformat()
        }
        
        return report


# ============================================================================
# GLOBAL SERVICE INSTANCE
# ============================================================================

evidence_vault = EvidenceVaultService()
